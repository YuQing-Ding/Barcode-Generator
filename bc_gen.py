"""

COPYRIGHT NOVA SCOTIA COMMUNITY COLLEGE - STRAIT AREA CAMPUS [ITGE]. ALL RIGHTS RESERVED.
PRODUCT MANAGER : DAVIS BOUDREAU
WRITTEN BY YUQING DING (SCOTT).
SPECIAL THANKS : CHATGPT (OPENAI).

"""

import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from PIL import Image, ImageDraw, ImageFont, ImageTk
import os
import csv
import random
import string
import openpyxl
from pyzbar.pyzbar import decode
import cv2
import barcode
from barcode.writer import ImageWriter
from datetime import datetime


class BarcodeGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Bar code generation system")
        self.root.geometry("500x400")

        # 创建标签和输入框
        self.room_label = tk.Label(self.root, text="Room number:")
        self.room_label.pack()
        self.room_entry = tk.Entry(self.root)
        self.room_entry.pack()

        self.type_label = tk.Label(self.root, text="Product Type:")
        self.type_label.pack()
        self.type_menu = tk.OptionMenu(self.root, tk.StringVar(), *self.get_csv_data("AssetType.csv"))
        self.type_menu.pack()

        self.brand_label = tk.Label(self.root, text="Product Brands:")
        self.brand_label.pack()
        self.brand_menu = tk.OptionMenu(self.root, tk.StringVar(), *self.get_csv_data("Brand.csv"))
        self.brand_menu.pack()

        self.location_label = tk.Label(self.root, text="Asset Location:")
        self.location_label.pack()
        self.location_menu = tk.OptionMenu(self.root, tk.StringVar(), *self.get_csv_data("Location.csv"))
        self.location_menu.pack()

        # 创建生成条形码按钮
        self.generate_button = tk.Button(self.root, text="Generate barcode", command=self.generate_barcode)
        self.generate_button.pack()

        # 创建扫描条形码按钮
        self.scan_button = tk.Button(self.root, text="Scan barcode", command=self.scan_barcode)
        self.scan_button.pack()

        # 创建摄像头显示框
        self.video_frame = tk.Label(self.root)
        self.video_frame.pack()

        # 用于存储已使用的编号
        self.used_numbers = {"room": set(), "type": set(), "brand": set(), "location": set(), "product": set()}
        self.load_used_numbers()

    def get_csv_data(self, filename):
        data = []
        with open(filename, "r") as file:
            reader = csv.reader(file)
            next(reader)  # 跳过标题行
            for row in reader:
                data.append(row[0])
        return data

    def generate_barcode(self):
        room_number = self.room_entry.get()
        product_type = self.type_menu.cget("text")
        product_brand = self.brand_menu.cget("text")
        asset_location = self.location_menu.cget("text")

        if not room_number or not product_type or not product_brand or not asset_location:
            messagebox.showerror("Error", "Please enter all information")
            return

        # 生成随机编号
        barcode_number = self.get_random_number_for_field("product")

        if barcode_number is None:
            return

        # 获取CSV行号
        type_index = self.get_csv_data("AssetType.csv").index(product_type) + 1
        brand_index = self.get_csv_data("Brand.csv").index(product_brand) + 1
        location_index = self.get_csv_data("Location.csv").index(asset_location) + 1

        # 生成条形码文本
        barcode_text = f"{room_number}-{type_index:03d}-{brand_index:03d}-{location_index:03d}-{barcode_number}"

        # 创建条形码图像
        barcode_image = self.generate_barcode_image(barcode_text)

        # 保存条形码图像
        save_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Image", "*.png")])
        if save_path:
            barcode_image.save(save_path)
            messagebox.showinfo("Saved", "Bar code image saved successfully")
        else:
            messagebox.showinfo("Save Cancelled", "Bar code image saving is cancelled")

        # 显示条形码图像
        barcode_image.show()

    def generate_barcode_image(self, barcode_text):
        # 创建条形码对象
        barcode_class = barcode.get_barcode_class("code128")
        barcode_object = barcode_class(barcode_text, writer=ImageWriter())

        # 设置条形码尺寸和边距
        options = {
            'module_width': 0.3,
            'module_height': 10.0,
            'quiet_zone': 1.0,
            'font_size': 12,
        }

        # 生成条形码图像
        barcode_image = barcode_object.render(options)

        # 在图像底部添加标注
        label_text = "TRAINING USE ONLY"
        label_font = ImageFont.truetype("arial.ttf", 20)
        label_width, label_height = label_font.getsize(label_text)
        image_width, image_height = barcode_image.size
        label_x = (image_width - label_width) // 2
        label_y = image_height - label_height - 10

        draw = ImageDraw.Draw(barcode_image)
        draw.text((label_x, label_y), label_text, font=label_font, fill="black")

        return barcode_image

    def scan_barcode(self):
        # 打开摄像头进行扫描
        video_capture = cv2.VideoCapture(0)

        def scan_loop():
            nonlocal video_capture

            ret, frame = video_capture.read()
            if not ret:
                messagebox.showerror("Error", "Cannot turn on the camera")
                return

            # 从图像中解码条形码
            barcodes = decode(frame)

            if len(barcodes) > 0:
                # 获取条形码数据
                barcode_data = barcodes[0].data.decode("utf-8")

                # 解析条形码数据
                barcode_parts = barcode_data.split('-')
                if len(barcode_parts) == 5:
                    room_number, type_index, brand_index, location_index, barcode_number = barcode_parts
                    product_type = self.get_csv_data("AssetType.csv")[int(type_index) - 1]
                    product_brand = self.get_csv_data("Brand.csv")[int(brand_index) - 1]
                    asset_location = self.get_csv_data("Location.csv")[int(location_index) - 1]

                    # 显示扫描结果
                    messagebox.showinfo("Scan Results",
                                        f"Room number: {room_number}\nProduct Type: {product_type}\nProduct Brands: {product_brand}\nAsset Location: {asset_location}\nBar Code: {barcode_number}")

                    # 写入Excel文件
                    self.write_to_excel(room_number, product_type, product_brand, asset_location, barcode_number)

                else:
                    messagebox.showerror("Error", "The scanned barcode is not in the correct format")

            # 在摄像头显示框中显示图像
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            image = Image.fromarray(frame)
            image = image.resize((400, 300))
            image = ImageTk.PhotoImage(image)
            self.video_frame.configure(image=image)
            self.video_frame.image = image  # 保持引用，避免被垃圾回收

            # 继续循环扫描
            self.root.after(1, scan_loop)

        scan_loop()

    def write_to_excel(self, room_number, product_type, product_brand, asset_location, barcode_number):
        if not os.path.exists("scan_results.xlsx"):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                ["Date of scanning", "Room number", "Product Type", "Product Brands", "Asset Location", "Product No."])
            workbook.save("scan_results.xlsx")

        workbook = openpyxl.load_workbook("scan_results.xlsx")
        sheet = workbook.active
        row_count = sheet.max_row + 1

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.cell(row=row_count, column=1, value=current_time)
        sheet.cell(row=row_count, column=2, value=room_number)
        sheet.cell(row=row_count, column=3, value=product_type)
        sheet.cell(row=row_count, column=4, value=product_brand)
        sheet.cell(row=row_count, column=5, value=asset_location)
        sheet.cell(row=row_count, column=6, value=barcode_number)

        workbook.save("scan_results.xlsx")
        messagebox.showinfo("Saved", "Scan results are saved to an Excel file")

    def get_random_number_for_field(self, field):
        used_numbers_file = f"used_numbers_{field}.txt"

        # 检查已使用号码文件是否存在，不存在则创建
        if not os.path.exists(used_numbers_file):
            open(used_numbers_file, 'w').close()

        # 读取已使用的号码集合
        used_numbers = self.used_numbers[field]

        # 生成随机编号，直到生成一个未使用的编号
        while True:
            barcode_number = ''.join(random.choice(string.digits) for _ in range(3))
            if barcode_number not in used_numbers:
                used_numbers.add(barcode_number)
                break
            elif len(used_numbers) == 999:
                messagebox.showerror("Error", f"All possible numbers have been used for {field}")
                return None

        # 将已使用的号码集合写入文件
        with open(used_numbers_file, 'w') as file:
            file.write('\n'.join(used_numbers))

        return barcode_number

    def load_used_numbers(self):
        for field in self.used_numbers.keys():
            used_numbers_file = f"used_numbers_{field}.txt"
            if os.path.exists(used_numbers_file):
                with open(used_numbers_file, 'r') as file:
                    self.used_numbers[field] = set(file.read().splitlines())


if __name__ == "__main__":
    root = tk.Tk()
    app = BarcodeGenerator(root)
    root.mainloop()
