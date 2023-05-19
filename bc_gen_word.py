"""

COPYRIGHT NOVA SCOTIA COMMUNITY COLLEGE - STRAIT AREA CAMPUS [ITGE]. ALL RIGHTS RESERVED.
PRODUCT MANAGER : DAVIS BOUDREAU
WRITTEN BY YUQING DING (SCOTT).
SPECIAL THANKS : CHATGPT (OPENAI).

"""

import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from PIL import Image, ImageDraw, ImageFont, ImageTk
import os
import csv
import random
import string
import openpyxl
from pyzbar.pyzbar import decode
import cv2
import barcode
from barcode.writer import ImageWriter
from datetime import datetime
import win32com.client as win32
from win32com.client import constants

class BarcodeGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Bar code generation system")
        self.root.geometry("500x400")
        # Initialize placeholder index
        self.placeholder_index = 1
        self.doc = None

        # Create labels and input fields
        self.room_label = tk.Label(self.root, text="Room number:")
        self.room_label.pack()
        self.room_entry = tk.Entry(self.root)
        self.room_entry.pack()

        self.type_label = tk.Label(self.root, text="Product Type:")
        self.type_label.pack()
        self.type_menu = tk.OptionMenu(self.root, tk.StringVar(), *self.get_csv_data("AssetType.csv"))
        self.type_menu.pack()

        self.brand_label = tk.Label(self.root, text="Product Brands:")
        self.brand_label.pack()
        self.brand_menu = tk.OptionMenu(self.root, tk.StringVar(), *self.get_csv_data("Brand.csv"))
        self.brand_menu.pack()

        self.location_label = tk.Label(self.root, text="Asset Location:")
        self.location_label.pack()
        self.location_menu = tk.OptionMenu(self.root, tk.StringVar(), *self.get_csv_data("Location.csv"))
        self.location_menu.pack()

        # Create Generate Barcode button
        self.generate_button = tk.Button(self.root, text="Generate barcode", command=self.generate_barcode)
        self.generate_button.pack()

        # Create scan barcode button
        self.scan_button = tk.Button(self.root, text="Scan barcode", command=self.scan_barcode)
        self.scan_button.pack()

        # Create camera display box
        self.video_frame = tk.Label(self.root)
        self.video_frame.pack()

        # Used to store used numbers
        self.used_numbers = {"room": set(), "type": set(), "brand": set(), "location": set(), "product": set()}
        self.load_used_numbers()

    def get_csv_data(self, filename):
        data = []
        with open(filename, "r") as file:
            reader = csv.reader(file)
            next(reader)  # Skip header line
            for row in reader:
                data.append(row[0])
        return data

    def generate_barcode(self):
        room_number = self.room_entry.get()
        product_type = self.type_menu.cget("text")
        product_brand = self.brand_menu.cget("text")
        asset_location = self.location_menu.cget("text")

        if not room_number or not product_type or not product_brand or not asset_location:
            messagebox.showerror("Error", "Please enter all information")
            return

        # Generate random number
        barcode_number = self.get_random_number_for_field("product")

        if barcode_number is None:
            return

        # Get CSV line number
        type_index = self.get_csv_data("AssetType.csv").index(product_type) + 1
        brand_index = self.get_csv_data("Brand.csv").index(product_brand) + 1
        location_index = self.get_csv_data("Location.csv").index(asset_location) + 1

        # Generate barcode text
        barcode_text = f"{room_number}-{type_index:03d}-{brand_index:03d}-{location_index:03d}-{barcode_number}"

        # Creating Barcode Images
        barcode_image = self.generate_barcode_image(barcode_text)

        # Save barcode images
        save_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Image", "*.png")])
        if save_path:
            barcode_image.save(save_path)
            messagebox.showinfo("Saved", "Bar code image saved successfully")
            if self.doc is None:
                word = win32.gencache.EnsureDispatch('Word.Application')
                self.doc = word.Documents.Open(os.path.abspath('Avery_old.doc'))
                word.Visible = False
            placeholder = f"PLACEHOLDER{self.placeholder_index}"
            self.replace_placeholder_with_image(self.doc, placeholder, save_path)
            self.doc.SaveAs(os.path.abspath('output.doc'))
            self.doc.Close(False)
            self.doc = None
            word.Quit()
            self.placeholder_index += 1
        else:
            messagebox.showinfo("Save Cancelled", "Bar code image saving is cancelled")

        # Display barcode images
        barcode_image.show()
    def replace_placeholder_with_image(self, doc, placeholder, image_path):
        # Find the placeholder
        for story in doc.StoryRanges:
            if story.Find.Execute(FindText=placeholder):
                # Delete the placeholder
                story.Delete()

                # Insert the image
                picture = story.InlineShapes.AddPicture(FileName=image_path)
                picture.LockAspectRatio = True
                picture.Width = 165  # Set width as needed

    def generate_barcode_image(self, barcode_text):
        # Creating Barcode Objects
        barcode_class = barcode.get_barcode_class("code128")
        barcode_object = barcode_class(barcode_text, writer=ImageWriter())

        # Set barcode size and margins
        options = {
            'module_width': 0.3,
            'module_height': 10.0,
            'quiet_zone': 1.0,
            'font_size': 12,
        }

        # Generate barcode images
        barcode_image = barcode_object.render(options)

        # Add annotations to the bottom of the image
        label_text = "TRAINING USE ONLY"
        label_font = ImageFont.truetype("arial.ttf", 20)
        label_width, label_height = label_font.getsize(label_text)
        image_width, image_height = barcode_image.size
        label_x = (image_width - label_width) // 2
        label_y = image_height - label_height - 10

        draw = ImageDraw.Draw(barcode_image)
        draw.text((label_x, label_y), label_text, font=label_font, fill="black")

        return barcode_image
    

    def scan_barcode(self):
        # Turn on the camera for scanning
        video_capture = cv2.VideoCapture(0)
        # Set the minimum size of the window
        self.root.minsize(500, 600)

        def scan_loop():
            nonlocal video_capture
            ret, frame = video_capture.read()
            # Get the frame rate of the camera
            fps = video_capture.get(cv2.CAP_PROP_FPS)
            # Adding the frame rate to the camera's video stream
            cv2.putText(frame, f"FPS: {fps}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)

            if not ret:
                messagebox.showerror("Error", "Cannot turn on the camera")
                return

            # Decoding barcodes from images
            barcodes = decode(frame)

            if len(barcodes) > 0:
                # Obtain barcode data
                barcode_data = barcodes[0].data.decode("utf-8")

                # Parsing barcode data
                barcode_parts = barcode_data.split('-')
                if len(barcode_parts) == 5:
                    room_number, type_index, brand_index, location_index, barcode_number = barcode_parts
                    product_type = self.get_csv_data("AssetType.csv")[int(type_index) - 1]
                    product_brand = self.get_csv_data("Brand.csv")[int(brand_index) - 1]
                    asset_location = self.get_csv_data("Location.csv")[int(location_index) - 1]

                    # Show scan results
                    messagebox.showinfo("Scan Results",
                                        f"Room number: {room_number}\nProduct Type: {product_type}\nProduct Brands: {product_brand}\nAsset Location: {asset_location}\nBar Code: {barcode_number}")

                    # Write to Excel file
                    self.write_to_excel(room_number, product_type, product_brand, asset_location, barcode_number)

                else:
                    messagebox.showerror("Error", "The scanned barcode is not in the correct format")

            # Displaying images in the camera display box
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            image = Image.fromarray(frame)
            image = image.resize((400, 300))
            image = ImageTk.PhotoImage(image)
            self.video_frame.configure(image=image)
            self.video_frame.image = image  # Keep reference to avoid garbage collection

            # Continue Loop Scan
            self.root.after(1, scan_loop)

        scan_loop()

    def write_to_excel(self, room_number, product_type, product_brand, asset_location, barcode_number):
        if not os.path.exists("scan_results.xlsx"):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                ["Date of scanning", "Room number", "Product Type", "Product Brands", "Asset Location", "Product No."])
            workbook.save("scan_results.xlsx")

        workbook = openpyxl.load_workbook("scan_results.xlsx")
        sheet = workbook.active
        row_count = sheet.max_row + 1

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.cell(row=row_count, column=1, value=current_time)
        sheet.cell(row=row_count, column=2, value=room_number)
        sheet.cell(row=row_count, column=3, value=product_type)
        sheet.cell(row=row_count, column=4, value=product_brand)
        sheet.cell(row=row_count, column=5, value=asset_location)
        sheet.cell(row=row_count, column=6, value=barcode_number)

        workbook.save("scan_results.xlsx")
        messagebox.showinfo("Saved", "Scan results are saved to an Excel file")

    def get_random_number_for_field(self, field):
        used_numbers_file = f"used_numbers_{field}.txt"

        # Check if the used number file exists, and create it if it does not.
        if not os.path.exists(used_numbers_file):
            open(used_numbers_file, 'w').close()

        # Retrieve the collection of used numbers
        used_numbers = self.used_numbers[field]

        # Generate random numbers until an unused number is generated
        while True:
            barcode_number = ''.join(random.choice(string.digits) for _ in range(3))
            if barcode_number not in used_numbers:
                used_numbers.add(barcode_number)
                break
            elif len(used_numbers) == 999:
                messagebox.showerror("Error", f"All possible numbers have been used for {field}")
                return None

        # Writing a collection of used numbers to a file
        with open(used_numbers_file, 'w') as file:
            file.write('\n'.join(used_numbers))

        return barcode_number

    def load_used_numbers(self):
        for field in self.used_numbers.keys():
            used_numbers_file = f"used_numbers_{field}.txt"
            if os.path.exists(used_numbers_file):
                with open(used_numbers_file, 'r') as file:
                    self.used_numbers[field] = set(file.read().splitlines())


if __name__ == "__main__":
    root = tk.Tk()
    app = BarcodeGenerator(root)
    root.mainloop()
